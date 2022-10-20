import calendar
import datetime

import django.core.exceptions
import environ
import openpyxl
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView
from django.db import connection
from django.db.models import Q, CharField, Value
from django.db.models.functions import Concat
from django.shortcuts import redirect
from django.shortcuts import render
from django.utils.datastructures import MultiValueDictKeyError

from diary import forms_choices as f
from diary.forms import ClientSearchForm, ClientDetailForm, ClientInputForm
from diary.forms import MemoForm, OfficerChangeForm, MainForm, PrintRequestForm, LoginForm
from diary.forms import WorkForm
from diary.models import ClientList
from diary.models import OfficerChange, Memo
from diary.models import Work
from diary.views import defs


# from django.template import Library
# register = Library()

class LoginViews(LoginView):
    template_name = "login.html"
    form_class = LoginForm


@login_required
def views_work(request, **kwargs):
    # 月日は2桁
    params = {'form': None, 'note': '', 'user': str(request.user),
              'year': datetime.date.today().strftime('%Y'),
              'month': datetime.date.today().strftime('%m'),
              'day': datetime.date.today().strftime('%d'),
              'day1': datetime.date.today().strftime('%d')}
    cur = connection.cursor()

    def get_last_date(dt):  # 月末が30か31かを判定している
        return dt.replace(day=calendar.monthrange(dt.year, dt.month)[1])

    if request.method == 'POST':
        form = WorkForm(request.POST)
        params['year'] = request.POST['year']
        params['month'] = request.POST['month'].zfill(2)
        if 'btn-c' in request.POST:  # 一日ごとの勤務時間を入力する処理
            params['day'] = request.POST['day'].zfill(2)
            params['in_time'] = '00:00' if request.POST['in_time'] == '' else request.POST['in_time']
            params['out_time'] = '00:00' if request.POST['out_time'] == '' else request.POST['out_time']
            params['rest_time'] = '00:00' if request.POST['rest_time'] == '' else request.POST['rest_time']
            params['note'] = request.POST['note']
            params['distance'] = 0 if request.POST['distance'] == '' else request.POST['distance']
            date = params['year'] + "-" + params['month'] + "-" + params['day']
            try:
                datetime.datetime.strptime(date, '%Y-%m-%d')
                if len(Work.objects.filter(name_id=params['user'], date=date)) == 0:  # 月に初めて入力するときに出勤表を作成
                    for_date = datetime.datetime.strptime(date, '%Y-%m-%d').replace(day=1)
                    for day in range(1, (get_last_date(for_date) - for_date).days + 2):
                        works = Work.objects.filter(name_id=params['user'],
                                                    date=datetime.date(for_date.year, for_date.month, day))
                        if works.exists() is False:
                            Work(name_id=params['user'], date=datetime.date(for_date.year, for_date.month, day),
                                         in_time='00:00', out_time='00:00', rest_time='00:00',
                                 note='', distance=0).save()
                u = Work.objects.get(name_id=params['user'], date=date)
                u.in_time = params['in_time']
                u.out_time = params['out_time']
                u.rest_time = params['rest_time']
                u.note = request.POST['note']
                u.distance = params['distance']
                u.save()
                params['check'] = "変更しました!"
            except (ValueError, django.core.exceptions.ValidationError):
                params['check'] = "日付が間違っている可能性があります。"
        elif 'btn-r2' in request.POST and params['user'] == f.users[1][0]:  # 年月で全員の出勤表を表示
            cur.execute(
                "select b.last_name || b.first_name, count(a.name_id), sum(a.out_time - a.in_time - a.rest_time), "
                "sum(a.distance), sum(a.distance * 20) from diary_work a join auth_user b on a.name_id=b.username "
                "where to_char(date,'MM')=%s and to_char(date,'YYYY')=%s "
                "and not in_time='00:00:00' group by b.last_name || b.first_name",
                (params['month'], params['year']))
            params['short_hyo'] = cur.fetchall()
            for i in range(0, len(params['short_hyo'])):  # 合計時間と燃料費を整える
                params['short_hyo'][i] = list(params['short_hyo'][i])
                params['short_hyo'][i][2] = round(params['short_hyo'][i][2] / datetime.timedelta(hours=1), 1)
                params['short_hyo'][i][4] = round(params['short_hyo'][i][4])
            cur.execute(
                "select a.id, a.date, a.in_time, a.out_time, a.rest_time, a.note, a.distance, a.name_id,"
                "a.out_time - a.in_time - a.rest_time, b.last_name, b.first_name from diary_work a "
                "join auth_user b on a.name_id=b.username where to_char(date,'YYYY')=%s and "
                "to_char(date,'MM')=%s order by name_id, date", (params['year'], params['month'])
            )
            params['long_hyo'] = cur.fetchall()
            params['check4'] = params['year'] + "年" + params['month'] + "月"
            params['check'] = '全体を表示します!'
        elif 'btn-r2' in request.POST and params['user'] != f.users[1][0]:
            params['check'] = '所長のみ操作可能です。'
        elif 'btn-check' in request.POST:
            params['check'] = "確認内容"
            a = Work.objects.filter(
                date__year=params['year'], date__month=params['month']
            ).order_by('name_id', 'date')
            for i in a:
                if i.sum[0] == "-":
                    params['check'] += "<br>" + i.name.last_name + i.name.first_name + str(i.date) + \
                                       "の合計がマイナスになっています。"
            if params['check'] == "確認内容":
                params['check'] = "エラーがありませんでした!"
        params['form'] = form
    else:
        params['date'] = datetime.date.today()
        params['year'] = params['date'].strftime('%Y')
        params['month'] = params['date'].strftime('%m')
        params['day'] = params['date'].strftime('%d')
        params['form'] = WorkForm(
            initial=dict(year=params['year'], month=params['month'], day=params['day'])
        )
    # 表示
    cur.execute(
        "select a.id, a.date, a.in_time, a.out_time, a.rest_time, a.note, a.distance, a.name_id, "
        "a.out_time - a.in_time - a.rest_time, b.last_name, b.first_name "
        "from diary_work a join auth_user b on a.name_id=b.username where name_id=%s and "
        "to_char(date,'YYYY')=%s and to_char(date,'MM')=%s order by date",
        (params['user'], params['year'], params['month'])
    )
    hyo_p = cur.fetchall()
    times = datetime.timedelta(hours=0)
    distances = 0.0
    for i in hyo_p:
        if i[8] is not None: times += i[8]
        if i[6] is not None: distances += i[6]
    # check2,3を修正する
    hyo_p_count = 0
    for i in hyo_p:
        if i[3] != datetime.time(hour=0, minute=0, second=0): hyo_p_count += 1
    params['check_dates'] = hyo_p_count
    params['check_times'] = round(times / datetime.timedelta(hours=1), 1)
    params['check_distances'] = distances
    params['check_gas'] = round(distances * 20)
    params['hyo'] = hyo_p
    return render(request, 'work.html', params)
def views_document(request):
    with open('./private_diary/diary/media/documents.txt', encoding="utf-8") as file:
        params = {'lists': file.readlines()}
        file.close()
    return render(request, 'document.html', params)


@login_required
def views_main(request, **kwargs):
    params = {'form': None, 'user': str(request.user)}  # 'ip': request.META.get('REMOTE_ADDR')
    initial_dict = dict(username=params['user'])
    # 認証をする
    if request.method == 'POST':
        form = MainForm(request.POST, initial=initial_dict)
        if 'btn-login' in request.POST:
            username = request.POST['username']
            env = environ.Env()
            password = env('USER_PASSWORD')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                params['check'] = "ログインしました！"
            else:
                params['check'] = "ユーザー名またはパスワードが間違っています。"
        elif 'btn-logout' in request.POST:
            logout(request)
            params['check'] = "ログアウトしました！"
        params['form'] = form
    else:
        if str(request.user) == 'AnonymousUser':
            params['check'] = 'ログインしてください'
        params['form'] = MainForm(initial=initial_dict)
    return render(request, 'main.html', params)


def view_officer_change(request):
    params = {'form': None}
    if request.method == 'POST':
        form = OfficerChangeForm(request.POST)
        if 'btn-c' in request.POST:
            params['number'] = defs.number_post(request.POST['number'], request.POST['number_choice'])
            params['year'] = request.POST['year']
            params['note'] = request.POST['note']
            try:
                u = OfficerChange.objects.get(k_number_id=params['number'], year=params['year'])
                u.note = params['note']
                params['check'] = "変更しました!"
            except OfficerChange.DoesNotExist:
                try:
                    c = ClientList.objects.get(k_number=params['number'], contract_cancellation="無")
                    OfficerChange(k_number_id=c.k_number, year=params['year'], note=params['note']).save()
                    params['check'] = "追加しました!"
                except ClientList.DoesNotExist:
                    params['check'] = "名簿に存在しません。"
        elif 'btn-d' in request.POST:
            params['number'] = defs.number_post(request.POST['number'], request.POST['number_choice'])
            params['year'] = request.POST['year']
            model = OfficerChange.objects.filter(k_number_id=params['number'], year=params['year'])
            model.delete()
            params['check'] = "内容を削除しました！"
        params['form'] = form
    else:
        params['form'] = OfficerChangeForm()
    params['hyo'] = OfficerChange.objects.select_related().all().order_by('k_number_id', 'year')
    params['form'].fields['number_choice'].choices = defs.number_choice()
    return render(request, 'officer_change.html', params)


def views_memo(request, **kwargs):
    params = {'name': '', 'form': None, 'result': '', 'user': request.user}
    initial_dict = dict(name=params['user'])
    if request.method == 'POST':
        form = MemoForm(request.POST, initial=initial_dict)
        params['name'] = request.POST['name']
        if 'btn-c' in request.POST:
            params['number'] = defs.number_post(request.POST['number'], request.POST['number_choice'])
            params['content'] = request.POST['content']
            try:
                Memo.objects.get(
                    k_number_id=params['number']
                ).content_memo = params['content']
                params['check'] = "更新しました!"
            except Memo.DoesNotExist:
                Memo(k_number_id=params['number'], content_memo=params['content']).save()
                params['check'] = "追加しました!"
        elif 'btn-d' in request.POST:
            params['number'] = defs.number_post(request.POST['number'], request.POST['number_choice'])
            Memo.objects.filter(k_number_id=params['number']).delete()
            params['check'] = "削除しました!"
        params['form'] = form
        params['hyo'] = Memo.objects.filter(
            Q(k_number__manager1=params['name']) | Q(k_number__manager2=params['name']),
            Q(k_number__add_memo='有'), Q(k_number__contract_cancellation='無')
        ).order_by('k_number_id')
    else:
        params['form'] = MemoForm(initial=initial_dict)
    params['form'].fields['number_choice'].choices = defs.number_choice()
    return render(request, 'memo.html', params)


def views_print_request(request):
    params = {'form': None}
    if request.method == 'POST':
        if 'btn-p' in request.POST:
            form = PrintRequestForm(request.POST)
            params['number'] = defs.number_post(request.POST['number'], request.POST['number_choice'])
            try:
                r = ClientList.objects.get(k_number=params['number'])
                excel_pass_name = '印刷依頼書EXCEL/' + r.k_name + ' 印刷依頼書.xlsx'
                wb = openpyxl.load_workbook('./private_diary/diary/media/印刷依頼書EXCEL/印刷依頼書テンプレート.xlsx')
                sheet = wb['Sheet1']
                sheet.cell(row=3, column=3).value = r.k_name
                if r.manager1 != '':
                    sheet.cell(row=3, column=9).value = r.manager1
                else:
                    sheet.cell(row=3, column=9).value = r.manager2
                wb.save('private_diary/diary/media/' + excel_pass_name)
                wb.close()
                params['check'] = 'エクセルを出力しました!(クリックするとダウンロードします)'
                url = str(request.META.get("HTTP_HOST"))
                params['file'] = 'http://' + url + '/diary/media/' + excel_pass_name
            except ClientList.DoesNotExist:
                params['check'] = "存在しません。"
            params['form'] = form
    else:
        params['form'] = PrintRequestForm()
    params['form'].fields['number_choice'].choices = defs.number_choice()
    return render(request, 'print_request.html', params)


def client_search(request):
    params = {'form': None}
    if request.method == "POST":
        form = ClientSearchForm(request.POST)
        if "btn-r" in request.POST:
            params['name'] = request.POST['name']
            params['hyo'] = ClientList.objects.filter(
                k_name__icontains=params['name'],).order_by('k_number')
            params['check'] = "表示しました!"
        params['form'] = form
    else:
        params['form'] = ClientSearchForm()
    return render(request, 'client_search.html', params)


def client_input(request):
    params = {'form': None, 'user': request.user, "ins_check": ClientList.objects.all().order_by("k_number")}
    form = ClientInputForm()
    # ここにチョイス書いてelseにchoice=を書く
    if request.user.is_authenticated:
        user = User.objects.get(username=str(request.user))
        user = user.last_name + " " + user.first_name
        form.fields['number_choice'].choices = ClientList.objects.select_related().filter(
            Q(manager1=user) | Q(manager2=user)).values_list(
            'k_number', Concat('k_number', Value(' '), 'k_name', output_field=CharField())).order_by('k_number')
    if 'btn-r' in request.GET:
        params['form'] = form
    elif 'btn-r1' in request.POST:
        params['type'] = request.POST['type']
        params['hyo'] = ClientList.objects.filter(type=params['type']).order_by('k_number')
        params['form'] = form
    else:
        params['form'] = form

    return render(request, 'client_input.html', params)


def views_client_detail(request):
    try:
        number = request.GET['k_number']
    except MultiValueDictKeyError:
        number = request.GET.get('number')

    if request.GET.get('number_choice') is not None:
        number = request.GET.get('number_choice')
    params = {'check': '詳細を表示します。', 'form': None, }
    if request.method == 'POST':
        form = ClientDetailForm(request.POST)
        params['form'] = form
        try:
            u = ClientList.objects.get(k_number=number)
            u.type = request.POST['type']
            u.k_ruby = request.POST['k_ruby']
            u.k_name = request.POST['k_name']
            u.company_name = request.POST['company_name']
            u.post_number = request.POST['post_number']
            u.prefecture = request.POST['prefecture']
            u.city = request.POST['city']
            u.tell_number = request.POST['tell_number']
            u.fax_number = request.POST['fax_number']
            u.representative_ruby = request.POST['representative_ruby']
            u.representative_mei = request.POST['representative_mei']
            u.representative_post_number = request.POST['representative_post_number']
            u.representative_address = request.POST['representative_address']
            u.manager1 = request.POST['manager1']
            u.manager2 = request.POST['manager2']
            u.manager_check = request.POST['manager_check']
            u.accounting_date = request.POST['accounting_date']
            u.tax_agency = request.POST['tax_agency']
            u.regional_tax_office = request.POST['regional_tax_office']
            u.government_office = request.POST['government_office']
            u.business_category = request.POST['business_category']
            u.report_category = request.POST['report_category']
            u.consumption_tax_category = request.POST['consumption_tax_category']
            u.withholding_tax_method = request.POST['withholding_tax_method']
            u.sales_management = request.POST['sales_management']
            u.report_schedule = request.POST['report_schedule']
            u.year_adjustment_check = request.POST['year_adjustment_check']
            u.consumption_tax_times = request.POST['consumption_tax_times']
            u.depreciable_assets = request.POST['depreciable_assets']
            u.extended_report = request.POST['extended_report']
            u.form_of_involvement = request.POST['form_of_involvement']
            u.k_corporate_tax = request.POST['k_corporate_tax']
            u.k_prefecture_tax = request.POST['k_prefecture_tax']
            u.k_city = request.POST['k_city']
            u.k_consumption = request.POST['k_consumption']
            u.k_income_tax = request.POST['k_income_tax']
            u.report_gift_tax = request.POST['report_gift_tax']
            u.add_memo = request.POST['add_memo']
            u.contract_cancellation = request.POST['contract_cancellation']
            u.involvement_date = request.POST['involvement_date']
            u.create_date = request.POST['create_date']
            u.last_update = datetime.date.today()
            u.t_filing = request.POST['t_filing']
            u.t_corporate_registration = request.POST['t_corporate_registration']
            u.t_important_points = request.POST['t_important_points']
            u.t_content = request.POST['t_content']
            u.t_start_date = request.POST['t_start_date']
            u.t_end_date = request.POST['t_end_date']
            u.t_history = request.POST['t_history']
            u.content_memo = request.POST['content_memo']
            u.electronic_report = request.POST['electronic_report']
            u.save()
            params['check'] = "変更しました！"
        except ClientList.DoesNotExist:
            params['check'] = "存在しません。"

        return render(request, 'client_detail.html', params)
    else:
        try:
            p = ClientList.objects.get(k_number=number)
            initial_dict = dict(type=p.type,
                                k_number=p.k_number,
                                k_ruby=p.k_ruby,
                                k_name=p.k_name,
                                company_name=p.company_name,
                                post_number=p.post_number,
                                prefecture=p.prefecture,
                                city=p.city,
                                tell_number=p.tell_number,
                                fax_number=p.fax_number,
                                representative_ruby=p.representative_ruby,
                                representative_mei=p.representative_mei,
                                representative_post_number=p.representative_post_number,
                                representative_address=p.representative_address,
                                manager1=p.manager1,
                                manager2=p.manager2,
                                manager_check=p.manager_check,
                                accounting_date=p.accounting_date,
                                tax_agency=p.tax_agency,
                                regional_tax_office=p.regional_tax_office,
                                government_office=p.government_office,
                                business_category=p.business_category,
                                report_category=p.report_category,
                                consumption_tax_category=p.consumption_tax_category,
                                withholding_tax_method=p.withholding_tax_method,
                                sales_management=p.sales_management,
                                report_schedule=p.report_schedule,
                                year_adjustment_check=p.year_adjustment_check,
                                consumption_tax_times=p.consumption_tax_times,
                                depreciable_assets=p.depreciable_assets,
                                extended_report=p.extended_report,
                                form_of_involvement=p.form_of_involvement,
                                k_corporate_tax=p.k_corporate_tax,
                                k_prefecture_tax=p.k_prefecture_tax,
                                k_city=p.k_city,
                                k_consumption=p.k_consumption,
                                k_income_tax=p.k_income_tax,
                                report_gift_tax=p.report_gift_tax,
                                add_memo=p.add_memo,
                                contract_cancellation=p.contract_cancellation,
                                involvement_date=p.involvement_date,
                                create_date=p.create_date,
                                last_update=p.last_update,
                                t_filing=p.t_filing,
                                t_corporate_registration=p.t_corporate_registration,
                                t_important_points=p.t_important_points,
                                t_content=p.t_content,
                                t_start_date=p.t_start_date,
                                t_end_date=p.t_end_date,
                                t_history=p.t_history,
                                content_memo=p.content_memo,
                                electronic_report=p.electronic_report
                                )
            params['check'] = "詳細を出力します！"
            params['form'] = ClientDetailForm(initial=initial_dict)
            return render(request, 'client_detail.html', params)
        except ClientList.DoesNotExist:
            params['check'] = "存在しません。"
            form = ClientInputForm(request.POST)
            params['form'] = form
            return redirect('../client_input/', params)